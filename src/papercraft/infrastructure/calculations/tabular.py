"""Trusted import of user tabular files into the shared Dataset model."""

from __future__ import annotations

import csv
import io
import math
import re
from collections.abc import Iterable, Sequence
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from papercraft.domain import Dataset, DatasetColumn, DataType, FactOrigin, Source


class TabularImportError(ValueError):
    pass


class TabularDatasetImporter:
    """Read CSV/XLSX without evaluating formulas or model-generated code."""

    def __init__(self, *, maximum_rows: int = 100_000, maximum_columns: int = 500) -> None:
        if maximum_rows < 1 or maximum_columns < 1:
            raise ValueError("tabular limits must be positive")
        self.maximum_rows = maximum_rows
        self.maximum_columns = maximum_columns

    def import_source(self, project_id: str, source: Source) -> list[Dataset]:
        path = Path(source.stored_path).resolve(strict=True)
        suffix = path.suffix.casefold()
        if suffix == ".csv":
            return [self._csv(project_id, source, path)]
        if suffix == ".xlsx":
            return self._xlsx(project_id, source, path)
        return []

    def _csv(self, project_id: str, source: Source, path: Path) -> Dataset:
        text = _decode(path.read_bytes())
        try:
            dialect = csv.Sniffer().sniff(text[:8192], delimiters=",;\t|")
        except csv.Error:
            dialect = csv.excel
        rows = self._bounded_rows(csv.reader(io.StringIO(text, newline=""), dialect))
        return _dataset(project_id, source, path.stem, rows, metadata={"delimiter": dialect.delimiter})

    def _xlsx(self, project_id: str, source: Source, path: Path) -> list[Dataset]:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:  # pragma: no cover
            raise TabularImportError("openpyxl is required to import XLSX") from exc
        try:
            workbook = load_workbook(path, read_only=True, data_only=True, keep_links=False)
        except Exception as exc:
            raise TabularImportError(f"Cannot read XLSX {path.name}: {exc}") from exc
        datasets: list[Dataset] = []
        try:
            for sheet in workbook.worksheets:
                rows = self._bounded_rows(sheet.iter_rows(values_only=True))
                if rows:
                    datasets.append(
                        _dataset(
                            project_id,
                            source,
                            f"{path.stem} — {sheet.title}",
                            rows,
                            metadata={"sheet": sheet.title},
                        )
                    )
        finally:
            workbook.close()
        return datasets

    def _bounded_rows(self, rows: Iterable[Sequence[Any]]) -> list[list[Any]]:
        result: list[list[Any]] = []
        for index, row in enumerate(rows):
            if index >= self.maximum_rows:
                break
            cleaned = [_json_value(value) for value in list(row)[: self.maximum_columns]]
            while cleaned and cleaned[-1] in {None, ""}:
                cleaned.pop()
            if cleaned:
                result.append(cleaned)
        return result


def _dataset(
    project_id: str,
    source: Source,
    name: str,
    raw_rows: list[list[Any]],
    *,
    metadata: dict[str, Any],
) -> Dataset:
    if not raw_rows:
        raise TabularImportError(f"{source.original_name} contains no data")
    width = max(len(row) for row in raw_rows)
    headers = _unique_headers(raw_rows[0], width)
    data_rows = raw_rows[1:]
    columns: list[DatasetColumn] = []
    column_types: list[DataType] = []
    for index, header in enumerate(headers):
        values = [row[index] if index < len(row) else None for row in data_rows]
        data_type = _infer_type(values)
        column_types.append(data_type)
        columns.append(
            DatasetColumn(
                name=header,
                data_type=data_type,
                nullable=any(value is None for value in values),
            )
        )
    rows = [
        {
            headers[index]: _coerce(
                row[index] if index < len(row) else None,
                column_types[index],
            )
            for index in range(width)
        }
        for row in data_rows
    ]
    return Dataset(
        project_id=project_id,
        name=name,
        columns=columns,
        rows=rows,
        origin=FactOrigin.USER,
        source_ids=[source.id],
        metadata={"imported": True, "source_sha256": source.sha256, **metadata},
    )


def _unique_headers(values: Sequence[Any], width: int) -> list[str]:
    headers: list[str] = []
    counts: dict[str, int] = {}
    for index in range(width):
        raw = values[index] if index < len(values) else None
        base = str(raw).strip() if raw not in {None, ""} else f"column_{index + 1}"
        base = base[:120]
        counts[base] = counts.get(base, 0) + 1
        headers.append(base if counts[base] == 1 else f"{base}_{counts[base]}")
    return headers


def _infer_type(values: Sequence[Any]) -> DataType:
    actual = [value for value in values if value is not None]
    if not actual:
        return DataType.STRING
    if all(isinstance(value, bool) for value in actual):
        return DataType.BOOLEAN
    if all(isinstance(value, int) and not isinstance(value, bool) for value in actual):
        return DataType.INTEGER
    if all(isinstance(value, (int, float)) and not isinstance(value, bool) for value in actual):
        return DataType.NUMBER
    if all(isinstance(value, datetime) for value in actual):
        return DataType.DATETIME
    if all(isinstance(value, date) for value in actual):
        return DataType.DATE
    textual = [value.strip() for value in actual if isinstance(value, str)]
    if len(textual) == len(actual):
        if all(re.fullmatch(r"[-+]?\d+", value) for value in textual):
            return DataType.INTEGER
        if all(_decimal_text(value) is not None for value in textual):
            return DataType.NUMBER
        if all(value.casefold() in {"true", "false"} for value in textual):
            return DataType.BOOLEAN
        if all(_iso_datetime(value) for value in textual):
            return DataType.DATETIME
        if all(_iso_date(value) for value in textual):
            return DataType.DATE
    return DataType.STRING


def _coerce(value: Any, data_type: DataType) -> Any:
    if value is None or not isinstance(value, str):
        return value
    cleaned = value.strip()
    if not cleaned:
        return None
    if data_type == DataType.INTEGER:
        return int(cleaned)
    if data_type == DataType.NUMBER:
        decimal = _decimal_text(cleaned)
        if decimal is None:
            return cleaned
        if decimal == decimal.to_integral_value():
            return int(decimal)
        converted = float(decimal)
        return converted if math.isfinite(converted) else cleaned
    if data_type == DataType.BOOLEAN:
        return cleaned.casefold() == "true"
    if data_type in {DataType.DATE, DataType.DATETIME}:
        return cleaned
    return value


def _decimal_text(value: str) -> Decimal | None:
    normalized = value.replace(" ", "").replace(",", ".")
    if re.fullmatch(r"[-+]?(?:\d+(?:\.\d*)?|\.\d+)", normalized) is None:
        return None
    try:
        converted = Decimal(normalized)
    except InvalidOperation:
        return None
    return converted if converted.is_finite() else None


def _iso_date(value: str) -> bool:
    try:
        date.fromisoformat(value)
    except ValueError:
        return False
    return True


def _iso_datetime(value: str) -> bool:
    if "T" not in value and " " not in value:
        return False
    try:
        datetime.fromisoformat(value)
    except ValueError:
        return False
    return True


def _decode(raw: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "cp1251"):
        try:
            return raw.decode(encoding)
        except UnicodeDecodeError:
            continue
    return raw.decode("utf-8", errors="replace")


def _json_value(value: Any) -> Any:
    if isinstance(value, str) and not value.strip():
        return None
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    return str(value)


__all__ = ["TabularDatasetImporter", "TabularImportError"]
